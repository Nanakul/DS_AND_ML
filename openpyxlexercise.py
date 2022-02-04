from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Create new worksheets
ws1 = wb.create_sheet('New Sheet')
ws2 = wb.create_sheet('First', 0)

# Alter title of current worksheet
ws.title = 'MySheet'

print(wb.sheetnames)

# Create a 2nd workbook and load an existing excel file.
wb2 = load_workbook('Regions.xlsx')
new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

# Access the value at cell A1
cell = active_sheet['A1']
print(cell.value)
active_sheet['A1'] = 'Region'

# Save modified excel sheet
wb2.save('modified_regions.xlsx')

# Learning how to navigate through data
# Create a range of cells to navigate
wb = load_workbook('regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']
col_range = ws['A': 'C']
print(col_range)

# Set a row range
row_range = ws[1:5]
print(row_range)

# Iterate through columns and rows.
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)

wb.save('regions.xlsx')
